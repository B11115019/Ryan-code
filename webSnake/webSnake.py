#載入 selenium 相關模組
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import time


#設定 Chrome Driver 的執行路徑
options = Options()
options.chrome_executable_path = "C:/Users/ryan/Desktop/chromdriver/chromedriver.exe"

#建立 Driver 物件 用程式執行瀏覽器操作
driver = webdriver.Chrome(options=options)

#連線到google新聞頁面 
driver.get("https://news.google.com/home?hl=zh-TW&gl=TW&ceid=TW:zh-Hant")
time.sleep(5)#等待五秒鐘

#捲動視窗並載入更多內容
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") #捲動到視窗底部
time.sleep(5)#等待五秒鐘

#取得焦點新聞內容
keyNews = driver.find_element(By.CLASS_NAME, "Ly25Ed")

#取得地方新聞內容
localNews = driver.find_element(By.CSS_SELECTOR, "[jsname=bYJ5Qe]")

#取得您的主題新聞內容
forMeNews = driver.find_element(By.CSS_SELECTOR, "[jsname=b90aOc]")

#匯入excel**************************
wb = Workbook()
ws = wb.active
#設定頁面名稱
ws.title = '今日新聞'

#做些字串處理
temp = keyNews.text.split('\n')
for i in temp:
    ws.append([i])
    #如果字體為特定字體，則對他做變色
    if(i == '焦點新聞'):
        i = ws[ws.max_row]
        for cell in i:
            cell.font = Font(color="FF0000")  # 將字體顏色設定為紅色

#將此字串匯入Excel
title1 = ['地方新聞']
ws.append(title1)
#將上述字串染色
title1_range = ws[ws.max_row]
for cell in title1_range:
    cell.font = Font(color="FF0000")  # 將字體顏色設定為紅色

#做些字串處理並匯入Excel
temp = localNews.text.split('\n')
for i in temp:
    ws.append([i])

#將此字串匯入Excel
title1 = ['您的主題']
ws.append(title1)
#將上述字串染色
title1_range = ws[ws.max_row]
for cell in title1_range:
    cell.font = Font(color="FF0000")  # 將字體顏色設定為紅色

#做些字串處理並匯入Excel
temp = forMeNews.text.split('\n')
for i in temp:
    ws.append([i])
    #如果字體為特定字體，則對他做變色
    if(i == '台灣' or i == "國際"or i == "商業"or i == "科學與科技"or i == "娛樂"or i == "體育"):
        i = ws[ws.max_row]
        for cell in i:
            cell.font = Font(color="0000FF")  # 將字體顏色設定為藍色

#儲存檔案
wb.save('daily news.xlsx')

driver.close()